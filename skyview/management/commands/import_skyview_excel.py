from datetime import date, datetime
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand

from skyview.models import FamilySite, Place, Video
from skyview.utils import convert_video_type, extract_youtube_id, shorten_title

FAMILY_SITE_SEED_DATA = [
    {
        "title": "English Study Site",
        "url": "https://tonykks.github.io/english-study-site/",
        "description": "영어 학습 자료 사이트",
        "order": 1,
    },
    {
        "title": "Hallim Youth English",
        "url": "https://tonykks.github.io/hallim-youth-english/",
        "description": "한림 청소년 영어 학습 사이트",
        "order": 2,
    },
]


class Command(BaseCommand):
    help = "Import Place and Video data from skyview.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="skyview.xlsx",
            help="Path to the Excel file (default: skyview.xlsx in project root)",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["file"])
        if not excel_path.is_absolute():
            excel_path = settings.BASE_DIR / excel_path

        if not excel_path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {excel_path}"))
            return

        workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        place_count = self._import_places(workbook["Places"])
        video_count = self._import_videos(workbook["Main"])
        family_site_count = self._import_family_sites()

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete: {place_count} places, "
                f"{video_count} videos, {family_site_count} family sites"
            )
        )

    def _import_places(self, sheet):
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            place_name = row[0]
            if not place_name:
                continue

            Place.objects.update_or_create(
                name=str(place_name).strip(),
                defaults={
                    "priority": self._to_int(row[1]),
                    "intro_url": self._to_str(row[2]) or "",
                    "latitude": self._to_float(row[3]),
                    "longitude": self._to_float(row[4]),
                    "is_active": True,
                },
            )
            count += 1
        return count

    def _import_videos(self, sheet):
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            youtube_url = self._to_str(row[2])
            if not youtube_url:
                continue

            raw_title = self._to_str(row[1]) or ""
            youtube_id = extract_youtube_id(youtube_url)
            lookup = {"youtube_id": youtube_id} if youtube_id else {"youtube_url": youtube_url}

            place = None
            place_name = self._to_str(row[4])
            if place_name:
                place = Place.objects.filter(name=place_name.strip()).first()

            defaults = {
                "place": place,
                "title": shorten_title(raw_title),
                "youtube_url": youtube_url,
                "youtube_id": youtube_id,
                "video_type": convert_video_type(row[3]),
                "published_date": self._to_date(row[0]),
                "is_active": True,
            }

            Video.objects.update_or_create(**lookup, defaults=defaults)
            count += 1
        return count

    def _import_family_sites(self):
        count = 0
        for item in FAMILY_SITE_SEED_DATA:
            FamilySite.objects.update_or_create(
                title=item["title"],
                defaults={
                    "url": item["url"],
                    "description": item["description"],
                    "order": item["order"],
                    "is_active": True,
                },
            )
            count += 1
        return count

    def _to_str(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def _to_int(self, value):
        if value is None or value == "":
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _to_float(self, value):
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _to_date(self, value):
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None
